from openpyxl import load_workbook
import openpyxl
from html.parser import HTMLParser
import boto3
import datetime
import requests
import lxml.html as html
from lxml.cssselect import CSSSelector
import smtplib
import io
import functools


##################################################################
## CONSTANTS #####################################################
##################################################################

s3_bucket_name = 'toll-reporting'
dynamo_table_name = "TollCredentials"
s3_file_name = 'Theracare_Expense_Report.xlsx'
cookie_url = 'https://csc.ntta.org/olcsc/AuthenticateUser.do'
data_url = 'https://csc.ntta.org/olcsc/DisplayHtmlTransactions.do?buttonClicked=Y'
table_headers = ['Transaction Date/Time', 'License Plate', 'Location', 'Transaction Type/Description', 'Amount']
exp_start_date_cell = 'K5'
exp_sum_cell = 'F10'

##################################################################
## FUNCS #########################################################
##################################################################

def get_credentials(email_address):
    dynamo = boto3.client('dynamodb')
    creds = dynamo.get_item(
        TableName=dynamo_table_name,
        Key={
            "EmailAddress": {
                "S":email_address
            }
        }
    )

    return {
        "username":creds["Item"]["Username"]["S"],
        "password":creds["Item"]["Password"]["S"]
    }

def get_expense():
    s3 = boto3.client('s3')
    report = s3.get_object(Bucket = s3_bucket_name, Key=s3_file_name)
    binary = report['Body'].read()
    return load_workbook(io.BytesIO(binary))


def get_dates():
    payload = {}
    today = datetime.date.today()
    payload['endDate'] = today.replace(day=1) - datetime.timedelta(days=1)
    payload['startDate'] = (payload['endDate'].replace(day=1))
    for key in payload:
        payload[key] = payload[key].strftime("%m/%d/%Y")
    return payload

def get_sum_data(data):
    amt_idx = 0
    for idx, header in enumerate(data):
        if header.lower() is 'amount':
            amt_idx = idx

    return functools.reduce((lambda a,b:a[amt_idx]+b[amt_idx]), data[1:])

#uses previous month
#input is credentials from dynamodb
#output is 2d array representing table to go in excel file
def get_transactions(creds):

    def grab_html(dates, creds):
        # get cookie
        payload = {
            'userName': creds['username'],
            'password': creds['password']
        }
        session = requests.Session()
        session.post(cookie_url, data=payload)

        # get data
        response = session.post(data_url, data=dates)

        # grab necessary table
        doc = html.fromstring(response.text)
        return doc.cssselect('table#record')[0]

    def map_table(table):
        # map to 2d array
        data = []

        # note down relevant indices, populate headers of data table
        headers = []
        column_indices = []
        for idx, header in enumerate(table.cssselect('th')):
            if header.text in table_headers:
                column_indices.append(idx)
                headers.append(header)
        data.append(headers)

        # grab relevant entries from each row
        for row in table.cssselect('tbody tr'):
            data_row = []
            entries = row.cssselect('td')
            for idx in column_indices:
                data_row.append(entries[idx])
            data.append(data_row)

        return data


    dates = get_dates()
    table = grab_html(dates, creds)
    return map_table(table)


#input is 2d data array
#output is newly created excel file
def build_data_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tolls'

    #nested loop to copy data into cells
    for x, row in enumerate(data):
        for y, entry in enumerate(row):
            ws.cell(row=x+1, column=y+1, value = entry)

    #sum amounts
    sum_col = 0
    for x, header in enumerate(data[0]):
        if header.lower() is 'amount':
            sum_col = x+1

    start_sum = ws.cell(row=1, column=sum_col)
    end_sum = ws.cell(row = data.len(), column=sum_col)
    sum_cell_val = '=SUM('+start_sum.coordinate+':'+end_sum.coordinate+')'
    ws.cell(row=data.len()+1, column=sum_col, value = sum_cell_val)

    return wb

def fill_expense(report, sum):
    ws = report.active
    start_date = get_dates()['startDate']
    ws[exp_start_date_cell].data_type = 'date'
    ws[exp_start_date_cell] = start_date
    ws[exp_sum_cell] = sum


## for testing
def save_locally(expense, tolls):
    expense.save('/tmp/expense.xlsx')
    tolls.save('/tmp/tolls.xlsx')


# def send_email(expense, tolls):


##################################################################
## lambda handler ################################################
##################################################################

# if __name__ == '__main__':
def lambda_handler(event, context):
    expense = get_expense()
    creds = get_credentials(event["requestContext"]["body"]["sender"])
    data = get_transactions(creds)
    tolls = build_data_excel(data)
    fill_expense(expense, get_sum_data(data))

    #save to s3 with date
    save_locally(expense, tolls)
    #send both to email